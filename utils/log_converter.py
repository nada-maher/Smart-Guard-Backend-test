"""
Log Conversion Utility
Converts CSV logs to XLSX format
"""
import csv
import os
from datetime import datetime

def csv_to_xlsx():
    """
    Convert inference_logs.csv to inference_logs.xlsx using openpyxl
    Requires: pip install openpyxl
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        csv_file = "inference_logs.csv"
        xlsx_file = "inference_logs.xlsx"
        
        if not os.path.exists(csv_file):
            print(f"CSV file '{csv_file}' not found")
            return False
        
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Inference Logs"
        
        # Read CSV and write to Excel
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            
            # Write header with styling
            header_row = next(reader)
            for col_num, header in enumerate(header_row, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write data rows
            for row_num, row in enumerate(reader, 2):
                for col_num, value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(xlsx_file)
        print(f"✓ Successfully converted CSV to XLSX: {xlsx_file}")
        return True
    
    except ImportError:
        print("Error: openpyxl is required. Install with: pip install openpyxl")
        return False
    except Exception as e:
        print(f"Error converting CSV to XLSX: {e}")
        return False

def org_csv_to_xlsx(organization_name):
    """
    Convert organization-specific CSV to XLSX format
    Requires: pip install openpyxl
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Sanitize organization name for file naming
        safe_org_name = "".join([c if c.isalnum() else "_" for c in organization_name])
        csv_file = f"logs/logs_{safe_org_name}.csv"
        xlsx_file = f"logs/logs_{safe_org_name}.xlsx"
        
        if not os.path.exists(csv_file):
            print(f"Organization CSV file '{csv_file}' not found")
            return False
        
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"{organization_name} Logs"
        
        # Read CSV and write to Excel
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            
            # Write header with styling
            header_row = next(reader)
            for col_num, header in enumerate(header_row, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write data rows
            for row_num, row in enumerate(reader, 2):
                for col_num, value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(xlsx_file)
        print(f"✓ Successfully converted organization CSV to XLSX: {xlsx_file}")
        return True
    
    except ImportError:
        print("Error: openpyxl is required. Install with: pip install openpyxl")
        return False
    except Exception as e:
        print(f"Error converting organization CSV to XLSX: {e}")
        return False

def csv_to_xlsx_pandas():
    """
    Alternative method using pandas
    Requires: pip install pandas openpyxl
    """
    try:
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment
        
        csv_file = "inference_logs.csv"
        xlsx_file = "inference_logs.xlsx"
        
        if not os.path.exists(csv_file):
            print(f"CSV file '{csv_file}' not found")
            return False
        
        # Read CSV
        df = pd.read_csv(csv_file)
        
        # Write to Excel with formatting
        with pd.ExcelWriter(xlsx_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Inference Logs', index=False)
            
            # Optional: Add formatting
            workbook = writer.book
            worksheet = writer.sheets['Inference Logs']
            
            # Format header
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        print(f"✓ Successfully converted CSV to XLSX: {xlsx_file}")
        return True
    
    except ImportError:
        print("Error: pandas and openpyxl are required. Install with: pip install pandas openpyxl")
        return False
    except Exception as e:
        print(f"Error converting CSV to XLSX: {e}")
        return False

if __name__ == "__main__":
    print("Converting inference logs from CSV to XLSX...")
    success = csv_to_xlsx()
    if not success:
        print("Trying alternative method with pandas...")
        csv_to_xlsx_pandas()
